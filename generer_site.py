

def liste_articles():
    import yaml

    with open('mkdocs.yml', 'r') as file:
        configuration = yaml.unsafe_load(file)

    liste = []
    for rubrique in configuration['nav'][2]["Leçons"]:
        for item in rubrique:
            for element in rubrique[item]:
                for k in element.keys():
                    entree = { 'titre' : k,  'fichier':element[k]}
                    liste.append(entree)
    return liste

def extraire_admonition(admonition, article):
    with open(article, 'r') as file:
        contenu = file.readlines()
    
    manuels = []
    admonition_tag = f"!!! {admonition}"

    entete_manuel = False
    for ligne in contenu:
        if admonition_tag in ligne:
            entete_manuel = True
        else:
            if entete_manuel and "](" in ligne:
                manuels.append(ligne.lstrip())
            else:
                entete_manuel = False

    return manuels

def generer_liste(admonition, entete, fichier_sortie):
    with open(fichier_sortie, 'w') as file:
        file.write(entete)

        articles = liste_articles()

        for article in articles:
            
            manuels = extraire_admonition(admonition, "wiki/" + article['fichier'])
            if manuels:
                file.write(f"## {article['titre']}")
                file.write("\n")

            for  manuel in manuels:
                file.write(manuel)

def generer_horaire():
    import openpyxl

    fichier = "horaire.xlsx"

    excel = openpyxl.load_workbook(fichier, data_only=True)

    feuille = excel["horaire"]

    with open("./wiki/horaire.md", "w") as f:
        # Écrire le titre de la page
        f.write("# Horaire du cours de développement web 3\n")

        # Écrire les entêtes du tableau de l'horaire
        entete = ""
        trait = ""
        for col in range(1, feuille.max_column + 1):
            entete += feuille.cell(row=1, column=col).value
            trait += "--"
            if col < feuille.max_column:
                entete += "|"
                trait += "|"
            else:
                entete += "\n"
                trait += "\n"

        f.write(entete)
        f.write(trait)

        for row in range(2, feuille.max_row + 1):
            ligne = ""
            for col in range(1, feuille.max_column + 1):
                valeur = feuille.cell(row=row, column=col).value
                if valeur is None:
                    valeur = ""
                else:
                    valeur = str(valeur)
                ligne += valeur
                if col < feuille.max_column:
                    ligne += "|"
                else:
                    ligne += "\n"

            f.write(ligne)


print("Générer la page des manuels")
generer_liste('manuel', "# Tous les manuels référencés dans le cours \n", "wiki/manuels.md")

print("Générer la page des CodeSandbox")
generer_liste('codesandbox', "# Tous les démos CodeSandbox référencés dans le cours \n", "wiki/codesandbox.md")

print("Générer la page de l'horaire")
generer_horaire()

